import openpyxl

def open_workbook(file):
    workbook = openpyxl.load_workbook(file)
    return workbook

def getRowCount(workbook,sheetName):    
    sheet = workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(workbook,sheetName):    
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)

def readData(workbook,sheetName,rownum,colnum):   
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum,column=colnum).value

def writeData(workbook,sheetName,rownum,colnum,data,filepath):    
    sheet = workbook[sheetName]
    sheet.cell(row=rownum,column=colnum).value = data
    if filepath != '':
        workbook.save(filepath)
        
    

